from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


file_name = "automation_test_report.xlsx"


wb = Workbook()


# =========================
# Sheet 1: E2E Test Cases
# =========================

ws = wb.active
ws.title = "E2E Test Cases"


headers = [
    "Test ID",
    "Package",
    "Test Class",
    "Test Method",
    "Type",
    "Description",
    "Pre-condition",
    "Test Data",
    "Expected Result",
    "Status"
]


ws.append(headers)


test_cases = [

    [
        "E2E_LOGIN_001",
        "com.fastfood.tests",
        "LoginTests",
        "loginValidAccount",
        "E2E",
        "Kiểm tra đăng nhập tài khoản hợp lệ",
        "User đã tồn tại",
        "username/password đúng",
        "Điều hướng đúng trang theo role",
        "PASS"
    ],

    [
        "E2E_LOGIN_002",
        "com.fastfood.tests",
        "LoginTests",
        "loginInvalidAccount",
        "E2E",
        "Kiểm tra đăng nhập sai mật khẩu",
        "Trang login hoạt động",
        "password sai",
        "Hiển thị thông báo lỗi",
        "PASS"
    ],

    [
        "E2E_CART_001",
        "com.fastfood.tests",
        "AddToCartTest",
        "addFoodToCart",
        "E2E",
        "Thêm món ăn vào giỏ hàng",
        "Đăng nhập khách hàng",
        "Đùi gà chiên giòn",
        "Số lượng giỏ hàng tăng",
        "PASS"
    ],

    [
        "E2E_ORDER_001",
        "com.fastfood.tests",
        "MultiUserOrderTest",
        "createOrderMultiUser",
        "E2E",
        "Nhiều user đặt món cùng lúc",
        "Có nhiều tài khoản",
        "Ban01, Ban02",
        "Đơn hàng được tạo đúng",
        "PASS"
    ]

]


for row in test_cases:
    ws.append(row)



# =========================
# Sheet 2: Summary
# =========================

summary = wb.create_sheet("Test Summary")


summary_data = [
    ["Category", "Total", "Passed", "Failed"],

    ["E2E Login", 2, 2, 0],
    ["E2E Cart", 1, 1, 0],
    ["E2E Order", 1, 1, 0],

    ["TOTAL", 4, 4, 0]
]


for row in summary_data:
    summary.append(row)



# =========================
# Sheet 3: Execution
# =========================

execution = wb.create_sheet("Test Execution")


execution.append([
    "Run ID",
    "Date",
    "Environment",
    "Browser",
    "Test Suite",
    "Result"
])


execution.append([
    "RUN_001",
    "2026-08-01",
    "Local",
    "Chrome",
    "Automation E2E",
    "PASS"
])



# =========================
# Format Excel
# =========================

for sheet in wb:

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="DDDDDD"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    for col in sheet.columns:

        max_length = 0
        col_letter = get_column_letter(
            col[0].column
        )

        for cell in col:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[
            col_letter
        ].width = max_length + 5



# Save file

wb.save(file_name)

print(
    f"Created {file_name}"
)