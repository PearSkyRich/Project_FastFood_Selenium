import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook


REPORT_FOLDER = "./target/surefire-reports"

EXCEL_FILE = "FastFood_Test_Case.xlsx"


# mở excel
wb = load_workbook(EXCEL_FILE)

ws = wb.active


# tạo map method -> status

status_map = {}


for file in os.listdir(REPORT_FOLDER):

    if file.endswith(".xml"):


        path = os.path.join(
            REPORT_FOLDER,
            file
        )


        tree = ET.parse(path)

        root = tree.getroot()


        for testcase in root.findall(".//testcase"):


            name = testcase.attrib["name"]


            failure = testcase.find("failure")


            if failure is None:

                status = "PASS"

            else:

                status = "FAIL"


            status_map[name] = status



# update excel

for row in range(2, ws.max_row + 1):

    method = ws.cell(
        row=row,
        column=4
    ).value


    if method in status_map:

        ws.cell(
            row=row,
            column=8
        ).value = status_map[method]



wb.save(EXCEL_FILE)


print("Update status hoàn thành")